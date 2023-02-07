from datetime import datetime
from datetime import timedelta
import imgkit
import sys

sys.path.append('../')

import utilities.file_utilities as file_utilities
import utilities.format_utilities as format_utilities
import utilities.dbutilities as dbutilities
import utilities.utilities as utilities
import utilities.column_names_utilities as cols
import openpyxl
import dataframe_image as dfi


from openpyxl.styles import Color,Font, Alignment, Border, Side, PatternFill,fills

from openpyxl.styles import colors
from openpyxl.cell import Cell

import pandas as pd

Total_Schools = "school_name"


working_schools = 'partial_yn'
working_status = ["1","2"]
groupby_index = ['district_name']
groupby_criteria = {"marked":"sum","partially_marked":"sum","unmarked":"sum",Total_Schools:"count"}

# Read the database connection credentials
credentials_dict = dbutilities.read_conn_credentials('db_credentials.json')

# Get the latest students and teachers count
df_report = dbutilities.fetch_data_as_df(credentials_dict, 'students_attendance_masterfile.sql')

#consider only the fully and partially working schools (yn is either 1 or 2) using isin.
df_filtered = utilities.filter_dataframe_not_in_column(df_report,working_schools,working_status)

#groupby district - keep this variable to add block if necessary - and sum marked,partially marked and unmarked and % calc
df_filtered = df_filtered.groupby(groupby_index).agg(groupby_criteria).reset_index()
df_filtered['% Marked Schools'],df_filtered['% Unmarked Schools'] =  df_filtered['marked']/df_filtered[Total_Schools],\
                                                                     df_filtered['unmarked']/df_filtered[Total_Schools]
df_filtered.rename(columns={'marked': 'Schools Marked', 'partially_marked': 'Schools Partially Marked','unmarked': 'Schools Unmarked',
                             'school_name':'Total Schools','district_name':'District'},
                                inplace=True)


#sortby highest unmarked
df_filtered=df_filtered.sort_values(['% Marked Schools'],ascending=True)

df_filtered.drop(columns=['% Unmarked Schools'], inplace = True)

function_dict = {"Schools Marked": "sum", "Schools Partially Marked":"sum","Schools Unmarked":"sum"
                 ,"Total Schools":"sum","% Marked Schools":"mean"}
df_filtered.loc[-1,:] = df_filtered.aggregate(function_dict)

df_filtered.at[-1, 'District'] = "Grand Total"

df_formatted = df_filtered.style.format("{:.2%}", subset ='% Marked Schools').set_properties(**{'text-align': 'center','border-color':'black','border-width':'0.1em'}).\
    background_gradient(cmap='RdYlGn',subset ='% Marked Schools')

df_formatted.set_table_styles([dict(selector='th', props=[('text-align', 'center','border-color','Black','border-width','0.1em')])])

datatoexcel = pd.ExcelWriter(r'C:\Users\Admin\Downloads\teachtest.xlsx', engine='xlsxwriter')

df_formatted.to_excel(datatoexcel, sheet_name='School Marking',index=False)
# save the excel
datatoexcel.save()


df_final = openpyxl.load_workbook('C:/Users/Admin/Downloads/teachtest.xlsx')
main_page = df_final['School Marking']
main_page.insert_rows(idx=0,amount=1)

prcnt_frmt = {'num_format': '0.00%'}
comp_schools_col_index = df_formatted.columns.get_loc('% Fully completed')
format_utilities.apply_frmt_cols(writer, 'Schools screening status', comp_schools_col_index, comp_schools_col_index,
                                 prcnt_frmt)










# First Row
col = main_page.column_dimensions['F']
col.number_format = 'Percent'
day1 =(datetime.today() - timedelta(days=2)).strftime('%m/%d/%Y')
day2 = (datetime.today() - timedelta(days=1)).strftime('%m/%d/%Y')
main_page['A1'].value = '% of Schools marking Teacher Attendance from {} to {}'.format(day1,day2)
font1 = Font(size=12, bold=True)
main_page['A1'].font = font1
main_page.merge_cells('A1:F1')
alignment=Alignment(horizontal='center',vertical='center')
main_page['A1'].alignment = alignment
bd=Side(border_style='thin',color='00000000')
border=Border(top=bd,left=bd,right=bd,bottom=bd)


df_final.save('C:/Users/Admin/Downloads/teachtest.xlsx')

