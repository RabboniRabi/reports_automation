import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook

# Calling the file that has all the sheets to be merged:
dir_containing_files = r'C:\Users\Admin\Downloads\test'

#Assigning a variable to call a workbook:
dest_wb = Workbook()

#Merging the sheets:
for root, dir, filenames in os.walk(dir_containing_files):
    for file in filenames:
        file_name = file.split('.')[0]
        # Absolute Path for Excel files:
        file_path = os.path.abspath(os.path.join(root, file))

        # Create new sheet in destination Workbook:
        dest_wb.create_sheet(file_name)
        dest_ws = dest_wb[file_name]

        # =====New Code====#

        # Read source data:
        source_wb = load_workbook(file_path)
        source_sheet = source_wb.active
        for row in source_sheet.rows:
            for cell in row:
                dest_ws[cell.coordinate] = cell.value

        # =================#

dest_wb.save(r"C:\Users\Admin\Downloads\test\tester.xlsx")

#Deleting the unnecessary worksheets:
dest_wb.remove(dest_wb['Sheet'])
dest_wb.remove(dest_wb['tester'])

#Assigning a variable for the Index Sheet:
sheet = dest_wb['master hyperlink']

#Assigning a variable for the last row to be filled:
max_row_num = (sheet.max_row)+1

#Defining a For loop to insert the '=HYPERLINK()' formula in each of the cells against the Index Items and formatting these cells too:
#range() creates the range from row 2 upto the last row that can be filled by added 1 (As the function does not include the last row itself
# the ".format() inserts the the row number in each of the {}s row by row :
for row_num in range(2, max_row_num):
    sheet['B{}'.format(row_num)] = '=HYPERLINK("[tester.xlsx]"&A{0}&"!A{0}","Click Here")'.format(row_num)
    sheet['B{}'.format(row_num)].font = Font(underline="single",bold=True,color='00000080',name='Cambria')

# Saving the file:
dest_wb.save(r"C:\Users\Admin\Downloads\test\tester.xlsx")





